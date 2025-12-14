from django.contrib import admin
from .models import Volunteer, Contact, Cause, Donate, Blog, Announcement
from django.http import HttpResponse
from django.urls import path
from django.shortcuts import redirect
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

class AnnouncementAdmin(admin.ModelAdmin):
    list_display = ('title', 'priority', 'is_active', 'created_at', 'end_date')
    list_filter = ('priority', 'is_active')
    search_fields = ('title', 'content')

class VolunteerAdmin(admin.ModelAdmin):
    list_display = ('name', 'email', 'subject', 'college_name', 'preprofessional')
    list_filter = ('college_name', 'preprofessional')
    search_fields = ('name', 'email', 'college_name', 'preprofessional')
    actions = ['export_to_excel']
    change_list_template = 'admin/volunteer_changelist.html'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-excel/', self.export_excel_view, name='volunteer_export_excel'),
        ]
        return custom_urls + urls

    def export_excel_view(self, request):
        return self.export_to_excel(request, Volunteer.objects.all())

    def export_to_excel(self, request, queryset):
        # Create a new workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Volunteers"

        # Define header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Add headers
        headers = ['ID', 'Name', 'Email', 'Subject', 'College/Company Name', 'Preprofessional', 'Message']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Add data
        for row_num, volunteer in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=volunteer.id)
            ws.cell(row=row_num, column=2, value=volunteer.name)
            ws.cell(row=row_num, column=3, value=volunteer.email)
            ws.cell(row=row_num, column=4, value=volunteer.subject)
            ws.cell(row=row_num, column=5, value=volunteer.college_name)
            ws.cell(row=row_num, column=6, value=volunteer.preprofessional)
            ws.cell(row=row_num, column=7, value=volunteer.message)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"volunteers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Save workbook to response
        wb.save(response)
        return response

    export_to_excel.short_description = "Export selected volunteers to Excel"

admin.site.register(Volunteer, VolunteerAdmin)
admin.site.register(Contact)
admin.site.register(Cause)
admin.site.register(Donate)
admin.site.register(Blog)
admin.site.register(Announcement, AnnouncementAdmin)